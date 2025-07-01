import datetime
import os.path
import random
import sys
import time
import traceback
import Api_engine
import math

import dbase_set
from time_tasks import time_for_now_tr, time_stamp_calculator, time_for_now

from PyQt6.QtCore import QRunnable, QThreadPool, pyqtSlot, QObject, pyqtSignal, QSize, QRect, Qt, QRegularExpression, \
    QTimer, QPoint, QRectF

from PyQt6.QtWidgets import QMainWindow, QApplication, QPushButton, QWidget, QSlider, QHBoxLayout, QProgressBar, \
    QVBoxLayout, QCheckBox, QLCDNumber, QSpinBox, QToolBar, QFormLayout, QLineEdit, QMessageBox, QGroupBox, QListWidget, \
    QListWidgetItem, QTabWidget, QFileDialog, QCalendarWidget, QCompleter, QButtonGroup, QLabel, QTableWidget, \
    QHeaderView, \
    QTableWidgetItem

from PyQt6.QtGui import QIcon, QAction, QIntValidator, QRegularExpressionValidator, QStandardItemModel, \
    QStandardItem, QDoubleValidator, QPainter, QColor, QFont, QPen, QFontMetrics

import data_manipulation as data_tasks

import pyqtgraph as pg

from dbase_set import push_order_items_to_dbase, push_orders_to_dbase, create_order_dbase, \
    create_comp_info_db, \
    push_scrap_datas_to_dbase, get_last_scrap_date, push_company_infos_to_dbase, get_seller_id_from_dbase, \
    get_comp_datas_from_dbase, get_selected_comps_datas_from_dbase, push_label_datas_to_dbase, get_ready_order_labels, \
    create_stock_datas_db, get_product_name_from_dbase, get_purchase_place_from_dbase, push_stock_datas_to_dbase, \
    push_match_datas_to_dbase, create_license_info_db, push_license_datas_to_dbase, get_license_status_from_dbase, \
    drop_license_info_db, update_label_printing_status, get_printed_order_labels, \
    create_label_datas_dbase, database_directory_name, get_waiting_order_items

import Api_engine

import license_check_tasks

from openpyxl import Workbook, load_workbook

import update_check_tasks
import subprocess

import pandas as pd

quit_turn = [1]

"""
Çıkma isteği sırası

"""


def scrap_data_from_api(scrap_time, comp_api_account_list: list, cntr: int, progress_status):
    """
    sipariş verilerini çekip,çektiklerini veritabanına yazdırır
    :param progress_status: ilerleme durumunu içinde çalıştığı fonksiyonun progress_status yapısına bildirir
    :param cntr: mevcut ilerleme durumunu takip edebilmek için sayaç
    :param comp_api_account_list: veri çekilmesi istenen hesapların api bilgilerini içeren bir liste
    :param scrap_time: epoch tipinden taramanın başlayacağı geçmiş tarih
    :return: sipariş verilerini çekip,çektiklerini veritabanına yazdırır
    """

    data_tasks.push_orders_to_tuple(scrap_time, comp_api_account_list)

    cntr += 1
    progress_status.emit(int(cntr * 100 / 3))

    push_order_items_to_dbase(data_tasks.order_items_list)
    cntr += 1
    progress_status.emit(int(cntr * 100 / 3))
    push_orders_to_dbase(data_tasks.order_datas_list)
    cntr += 1
    progress_status.emit(int(cntr * 100 / 3))


def run_all_pull_tasks(comp_api_account_list: list, progress_callback=None):
    """
    thread içerisinde çalıştırılacak olan fonksiyonların birlikte toplanmış hali
    :param comp_api_account_list: veri çekilmesi istenen hesapların api bilgilerini içeren bir liste
    :param progress_callback: işlemlerin sırası ile ilgili sinyal alır
    :return:thread içerisinde yapılacak işleri toplar ve çalıştırıp işlem durumunu sinyal ile döndürür
    """

    while True:

        if window.auto_check_box.isChecked():

            scrap_data_from_api("auto", comp_api_account_list, 0, progress_callback)

        else:

            scrap_data_from_api(time_for_now() - time_stamp_calculator(window.spin_box.value()),
                                comp_api_account_list, 0,
                                progress_callback)

        push_scrap_datas_to_dbase(time_for_now())

        data_tasks.push_labels_to_tuple()

        push_label_datas_to_dbase(data_tasks.label_datas_list)

        window.change_print_wait_button_activate()
        time.sleep(window.slider.value())

        if window.slider.value() == 0 or quit_turn[0] == 0:
            break


def run_all_profit_calc_tasks(excel_file_path, other_prices, progress_callback: None):
    profit = data_tasks.calculate_profit_with_excel(excel_file_path, other_prices, progress_callback=progress_callback)
    window.change_p_bar_text_profit(str(profit[0]) + "₺")
    if round(profit[0]) > 0:
        window.circ_pbar.data_label.setStyleSheet("color:green")
    else:
        window.circ_pbar.data_label.setStyleSheet("color:red")

    window.profit_label.setText(str(profit[0]))
    window.total_product_price_label.setText(str(round(profit[2])))
    window.total_commission_label.setText(str(round(profit[3])))
    window.total_tax_label.setText(str(round(profit[4])))
    window.total_cargo_label.setText(str(round(profit[5])))
    window.total_otp_label.setText(str(round(profit[6])))
    window.total_sell_price_label.setText(str(round(profit[7])))

    window.d_item_pos.clear()
    window.d_item_neg.clear()
    data_x_pos = []
    data_y_pos = []

    data_x_neg = []
    data_y_neg = []

    for data in profit[1].iterrows():
        if data[1][1] > 0:
            data_x_pos.append(data[0])
            data_y_pos.append(data[1][1])
        else:
            data_x_neg.append(data[0])
            data_y_neg.append(data[1][1])

    window.d_item_pos.setData(data_x_pos, data_y_pos)
    window.d_item_neg.setData(data_x_neg, data_y_neg)

    window.d_item_pos.setPen(pg.mkPen(color='green', width=2))
    window.d_item_neg.setPen(pg.mkPen(color='red', width=2))

    window.plot_widget.addItem(window.d_item_pos)
    window.plot_widget.addItem(window.d_item_neg)
    window.plot_widget.setXRange(0, len(data_x_pos) + len(data_x_neg) * 2)


class WorkerSignals(QObject):
    """
    thread içerisinde iletimleri sağlayan sinyallerin sınıfı
    """
    finished = pyqtSignal()
    error = pyqtSignal(tuple)
    result = pyqtSignal(object)
    progress = pyqtSignal(int)


class Worker(QRunnable):
    """
    thread sınıfıdır.herhangi bir thread içerisine yerleştirilirse çalışacaktır
    run:thread içerisinde çalışacak olan fonksiyon
    """

    def __init__(self, fn, *args, **kwargs):
        super().__init__()

        self.signals = WorkerSignals()

        self.fn = fn
        self.args = args
        self.kwargs = kwargs

        self.kwargs['progress_callback'] = self.signals.progress

    @pyqtSlot()
    def run(self):
        try:
            result = self.fn(*self.args, **self.kwargs)

        except:
            exc_type, value = sys.exc_info()[:2]
            self.signals.error.emit((exc_type, value, traceback.format_exc()))

        else:
            self.signals.result.emit(result)
        finally:
            self.signals.finished.emit()


class LicenseValidationWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.validate_button = QPushButton("Doğrula")
        self.disconnect_button = QPushButton("Bağlantıyı Kes")

        self.validate_button.setVisible(False)
        self.disconnect_button.setVisible(False)

        self.token_input = QLineEdit()
        self.mac_id_input = QLineEdit()
        self.key_input = QLineEdit()

        self.result_label = QLabel("")
        self.key_label = QLabel("Lisans Anahtarı:")
        self.mac_id_label = QLabel("MAC ID:")
        self.token_label = QLabel("Token:")

        self.token_input.setVisible(False)
        self.mac_id_input.setVisible(False)
        self.key_input.setVisible(False)
        self.key_label.setVisible(False)
        self.mac_id_label.setVisible(False)
        self.token_label.setVisible(False)

        self.option_buttons_layout = QHBoxLayout()
        self.validate_options_button = QPushButton("Doğrulama İşlemleri")
        self.disconnect_options_button = QPushButton("Lisans Bağlantısını Kes İşlemleri")
        self.option_buttons_layout.addWidget(self.validate_options_button)
        self.option_buttons_layout.addWidget(self.disconnect_options_button)

        self.table_widget_label = QLabel("Lisans ile etkinleşmiş cihazlar")
        self.table_widget_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.table_widget = QTableWidget()
        self.setWindowTitle("Lisans Doğrulama")
        self.setGeometry(100, 100, 600, 400)

        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        self.validate_button.setEnabled(False)
        self.key_input.textChanged.connect(self.change_validate_button_activate_status)

        self.mac_id_input.textChanged.connect(self.change_dsc_button_activate_status)

        self.disconnect_button.setEnabled(False)

        form_layout = QFormLayout()

        self.mac_id_input.setPlaceholderText("Eğer bir lisans ile makine arasındaki bağlantıyı keseceksen doldur!")
        self.key_input.setPlaceholderText("Bir lisans anahtarı gir")

        self.key_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        form_layout.addRow(self.key_label, self.key_input)
        form_layout.addRow(self.mac_id_label, self.mac_id_input)
        form_layout.addRow(self.token_label, self.token_input)

        layout.addLayout(self.option_buttons_layout)
        layout.addLayout(form_layout)

        self.validate_options_button.clicked.connect(self.show_validate_options)
        self.disconnect_options_button.clicked.connect(self.show_disconnect_options)

        layout.addWidget(self.validate_button)
        layout.addWidget(self.disconnect_button)

        layout.addWidget(self.result_label)

        layout.addWidget(self.table_widget_label)  # Add the label above the table
        self.table_widget.setColumnCount(4)
        self.table_widget.setHorizontalHeaderLabels(["MAC Adresi", "Bilgisayar İsmi", "İşletim Sistemi", "Makine İD"])
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table_widget)

        self.setLayout(layout)

    def show_validate_options(self):
        self.token_input.setVisible(True)
        self.key_input.setVisible(True)
        self.key_label.setVisible(True)
        self.token_label.setVisible(True)
        self.validate_button.setVisible(True)

        self.mac_id_input.setVisible(False)
        self.mac_id_label.setVisible(False)
        self.disconnect_button.setVisible(False)

    def show_disconnect_options(self):
        self.token_input.setVisible(True)
        self.mac_id_input.setVisible(True)
        self.mac_id_label.setVisible(True)
        self.token_label.setVisible(True)
        self.disconnect_button.setVisible(True)

        self.key_input.setVisible(False)
        self.key_label.setVisible(False)
        self.validate_button.setVisible(False)

    def change_validate_button_activate_status(self):
        if len(self.key_input.text()) > 0:
            self.validate_button.setEnabled(True)
        else:
            self.validate_button.setEnabled(False)

    def change_dsc_button_activate_status(self):
        if len(self.mac_id_input.text()) > 0:
            self.disconnect_button.setEnabled(True)
        else:
            self.disconnect_button.setEnabled(False)


class CircularProgress(QWidget):
    def __init__(self):
        super().__init__()
        self.setMinimumSize(200, 200)
        self.progress = 0
        self.label_text = ""

        self.data_label = QLabel(self)  # Etiketi oluştur
        self.data_label.setFont(QFont("Arial", 30))
        self.painter = QPainter(self)

    def paintEvent(self, event):
        self.painter.begin(self)
        self.painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.painter.setFont(QFont('Arial', 10))

        # Center coordinates and radius of the circle
        center = self.rect().center()
        radius = min(self.width(), self.height()) / 3

        # Draw the outer circle as progress bar
        self.painter.setPen(QPen(QColor('#DDDDDD'), 25))
        self.painter.drawEllipse(center, radius, radius)

        # Calculate the progress arc
        progress_pen = QPen(QColor('#0077FF'), 25)
        progress_pen.setCapStyle(Qt.PenCapStyle.RoundCap)
        self.painter.setPen(progress_pen)
        angle = -self.progress * 3.6  # Convert progress to angle (0-360)
        self.painter.drawArc(center.x() - radius, center.y() - radius, radius * 2, radius * 2, 90 * 16, angle * 16)

        self.fitLabel()
        self.painter.end()

    def setProgress(self, value):
        self.progress = value
        self.update()

    def add_text_to_center(self, text: str):
        self.data_label.setText(text)
        self.label_text = text
        self.fitLabel()

    def fitLabel(self):
        self.data_label.setFont(QFont("Arial", 30))
        center = self.rect().center()
        radius = min(self.width(), self.height()) / 3

        # Hesaplamalar
        font = self.data_label.font()
        metrics = QFontMetrics(font)
        while metrics.horizontalAdvance(self.label_text) > radius * 1.5:
            font.setPointSize(font.pointSize() - 0.5)
            metrics = QFontMetrics(font)

        self.data_label.setFont(font)
        label_width = metrics.horizontalAdvance(self.label_text)
        self.data_label.setGeometry(center.x() - label_width // 2, center.y() - metrics.height() // 2, label_width,
                                    metrics.height())


class CurrencyTextEdit(QLineEdit):
    def __init__(self):
        super().__init__()

        self.setValidator(QIntValidator(0, 999999999, self))  # Sayı aralığı
        self.setText("₺0")  # Varsayılan değer

        self.textChanged.connect(self.fix_seperator_char)
        self.textChanged.connect(self.fix_default_value)
        self.textChanged.connect(self.fix_numeric_area)

    def fix_seperator_char(self):
        text = self.text().replace(",", ".")
        self.setText(text)

    def fix_default_value(self):
        if len(self.text()) < 2:
            self.setText("₺0")
        elif len(self.text()) == 3 and self.text()[0:3] != "₺0." and self.text()[2] != "0":
            text = self.text().replace("0", "")
            self.setText(text)
        else:
            pass

    def fix_numeric_area(self):
        if self.text()[1:].isnumeric() and "." not in self.text()[1:]:
            pass
        else:
            for char in self.text():
                if char in "₺.," or char.isnumeric():
                    pass
                else:
                    fixed_from_numerics = self.text().replace("{}".format(char), "")
                    self.setText(fixed_from_numerics)


class CompAddWindow(QWidget):
    """
    This "window" is a QWidget. If it has no parent, it
    will appear as a free-floating window as we want.
    """

    def __init__(self):
        super().__init__()
        self.setWindowTitle("ŞİRKET EKLEME ALANI")

        self.setWindowIcon(QIcon("images/sniper_icon.png"))

        self.save_button = QPushButton()
        self.save_button.setFixedSize(32, 32)
        self.save_button.setIcon(QIcon("images/save_button_32.png"))
        self.save_button.setIconSize(self.save_button.size())
        self.save_button.setFlat(True)
        self.save_button.setStyleSheet('''
                            QPushButton {
                                border: 0;
                                color: lightGray;
                                font-size: 1px;
                            }
                    ''')
        self.save_button.clicked.connect(self.save_button_action)

        self.setFixedSize(350, 150)

        self.layout_form = QFormLayout()
        layout_vertical = QVBoxLayout()

        self.int_validator_mask = QRegularExpressionValidator(QRegularExpression("[0-9]+"))

        self.comp_name = QLineEdit(self)
        self.api_key = QLineEdit(self)
        self.api_secret = QLineEdit(self)

        self.seller_id = QLineEdit(self)
        self.seller_id.setValidator(self.int_validator_mask)

        self.message_box = QMessageBox(self)

        self.layout_form.addRow("ŞİRKET ADI", self.comp_name)
        self.layout_form.addRow("APİ KEY", self.api_key)
        self.layout_form.addRow("APİ SECRET", self.api_secret)
        self.layout_form.addRow("SELLER İD", self.seller_id)

        self.layout_form.setContentsMargins(0, 0, 0, 12)

        layout_vertical.addLayout(self.layout_form)
        layout_vertical.addWidget(self.save_button)
        self.setLayout(layout_vertical)

    def save_button_action(self):
        try:

            saving_test_result_seller_id = get_seller_id_from_dbase(int(self.seller_id.text()))
            if len(saving_test_result_seller_id) == 1:
                self.message_box.setText("Zaten böyle bir kayıt var")
                self.message_box.show()


            else:

                push_company_infos_to_dbase(int(self.seller_id.text()), self.api_key.text(),
                                            self.api_secret.text(), self.comp_name.text())

                saving_test_result_seller_id = get_seller_id_from_dbase(int(self.seller_id.text()))
                if len(saving_test_result_seller_id) == 1:
                    self.message_box.setText("Başarıyla Kaydedildi")
                    self.message_box.show()
                    window.list_widget.clear()
                    window.add_cbox_to_list_widget()
                else:
                    self.message_box.setText("Kayıt sırasında hata oluştu")
                    self.message_box.show()
        except Exception as e:
            print(e)


class PrintedOrderWindow(QWidget):
    def __init__(self, printed_orders):
        super().__init__()

        self.excel_file_path_for_extract_orders = None

        self.table = QTableWidget()
        self.filter_widgets = []  # Filtre widget'larını saklamak için liste

        self.main_layout = QHBoxLayout()

        self.ver_layout_1 = QVBoxLayout()
        self.ver_layout_2 = QVBoxLayout()

        self.send_to_waiting_button = QPushButton("Bekleyene Gönder")
        self.select_all_button = QPushButton("TÜMÜNÜ SEÇ")
        self.deselect_all_button = QPushButton("TÜM SEÇİMİ KALDIR")

        self.column_list = ["Sipariş Numarası", "Kargo Takip Numarası", "Kargo Firması", "Müşteri Adı", "Soyadı",
                            "1. ürün", "adet", "2. ürün", "adet", "3. ürün", "adet", "4. ürün", "adet", "5. ürün",
                            "adet",
                            "6. ürün", "adet", "7. ürün", "adet", "8. ürün", "adet", "etiket numarası",
                            "kargo takip numarası (sayısal)", "full adres", "SEÇ"]
        self.printed_orders = printed_orders
        self.initUI()

    def initUI(self):
        self.setWindowTitle("YAZDIRILAN SİPARİŞLER")
        self.showMaximized()

        self.select_all_button.clicked.connect(self.select_all_cbox)
        self.deselect_all_button.clicked.connect(self.deselect_all_cbox)

        self.send_to_waiting_button.setEnabled(False)  # Başlangıçta butonu pasif yap

        self.table.setColumnCount(len(self.column_list))
        self.table.setHorizontalHeaderLabels(self.column_list)

        self.send_to_waiting_button.clicked.connect(self.send_to_waiting_orders)

        if os.path.exists(os.path.join(database_directory_name, "label_datas.db")):
            row_len = self.printed_orders.shape[0]
            self.table.setRowCount(row_len)
            check_box_location = len(self.column_list) - 1

            # Filtre widget'ları ekleyin
            filter_columns = [0, 1, 2, 3,
                              4]  # Sipariş Numarası, Kargo Takip Numarası, Kargo Firması, Müşteri Adı, Soyadı
            header_layout = QHBoxLayout()
            for col in range(self.table.columnCount()):
                if col in filter_columns:
                    filter_widget = QLineEdit()
                    filter_widget.setPlaceholderText(f"Filtrele {self.column_list[col]}")
                    filter_widget.textChanged.connect(self.apply_filter)
                    header_layout.addWidget(filter_widget)
                    self.filter_widgets.append((col, filter_widget))
                else:
                    spacer = QLabel("")
                    header_layout.addWidget(spacer)

            self.ver_layout_1.addLayout(header_layout)

            for row in range(row_len):
                checkbox = QCheckBox()
                checkbox.stateChanged.connect(self.update_send_button_state)  # Checkbox durum değişikliğine bağla
                self.table.setCellWidget(row, check_box_location, checkbox)

            self.table.resizeColumnsToContents()
            self.ver_layout_1.addWidget(self.table)
            self.ver_layout_2.addWidget(self.send_to_waiting_button)
            self.ver_layout_2.addWidget(self.select_all_button)
            self.ver_layout_2.addWidget(self.deselect_all_button)

            self.main_layout.addLayout(self.ver_layout_1)
            self.main_layout.addLayout(self.ver_layout_2)

            self.setLayout(self.main_layout)

    def update_send_button_state(self):
        """Checkboxlardan herhangi biri seçiliyse butonu aktif yap, değilse pasif yap."""
        any_checked = any(self.table.cellWidget(row, len(self.column_list) - 1).isChecked()
                          for row in range(self.table.rowCount()))
        self.send_to_waiting_button.setEnabled(any_checked)

    def load_data_to_table(self, progress_callback=None):
        df = self.printed_orders

        for row in range(df.shape[0]):
            for col in range(df.shape[1] - 1):
                self.table.setItem(row, col, QTableWidgetItem(str(df.iloc[row, col])))

    def send_to_waiting_orders(self):
        try:
            for row in range(self.table.rowCount()):
                checkbox = self.table.cellWidget(row, len(self.column_list) - 1)

                if checkbox and checkbox.isChecked():
                    selected_order_number = self.table.item(row, self.column_list.index("Sipariş Numarası"))
                    update_label_printing_status(selected_order_number.text(), "False")

            QMessageBox.information(self, 'Başarılı',
                                    'Siparişlerin statüsü bekleyene çekildi',
                                    QMessageBox.StandardButton.Ok)

        except Exception as save_except:
            QMessageBox.critical(self, 'Hata', 'sipariş statü değiştirme hatası:\n{}'.format(str(save_except)),
                                 QMessageBox.StandardButton.Ok)

    def select_all_cbox(self):
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row):  # Sadece görünen satırlar
                checkbox = self.table.cellWidget(row, len(self.column_list) - 1)
                checkbox.setCheckState(Qt.CheckState.Checked)
        self.update_send_button_state()

    def deselect_all_cbox(self):
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row):  # Sadece görünen satırlar
                checkbox = self.table.cellWidget(row, len(self.column_list) - 1)
                checkbox.setCheckState(Qt.CheckState.Unchecked)
        self.update_send_button_state()

    def apply_filter(self):
        """Filtreleme işlemi."""
        for row in range(self.table.rowCount()):
            show_row = True
            for col, filter_widget in self.filter_widgets:
                filter_text = filter_widget.text().lower()
                item_text = self.table.item(row, col).text().lower() if self.table.item(row, col) else ""
                if filter_text not in item_text:
                    show_row = False
                    break
            self.table.setRowHidden(row, not show_row)


class WaitingOrderWindow(QWidget):
    def __init__(self, waiting_orders):
        super().__init__()
        self.extract_xlsx_icon = QIcon("images/xlsx_extract.ico")

        self.excel_file_path_for_extract_orders = None

        self.table = QTableWidget()
        self.filter_widgets = []

        self.main_layout = QHBoxLayout()

        self.ver_layout_1 = QVBoxLayout()
        self.ver_layout_2 = QVBoxLayout()

        self.extract_orders_button = QPushButton("Excel çıkart")
        self.select_all_button = QPushButton("TÜMÜNÜ SEÇ")
        self.deselect_all_button = QPushButton("TÜM SEÇİMİ KALDIR")

        self.column_list = ["Sipariş Numarası", "Kargo Takip Numarası", "Kargo Firması", "Müşteri Adı", "Soyadı",
                            "1. ürün", "adet", "2. ürün", "adet", "3. ürün", "adet", "4. ürün", "adet", "5. ürün",
                            "adet",
                            "6. ürün", "adet", "7. ürün", "adet", "8. ürün", "adet", "etiket numarası",
                            "kargo takip numarası (sayısal)", "full adres", "SEÇ"]
        self.excel_column_list = ["orderNumber", "cargoTrackingNumber", "cargoProviderName", "customerName",
                                  "customerSurname",
                                  "leftFirstProd", "leftFirstQuantity", "leftSecondProd", "leftSecondQuantity",
                                  "leftThirdProd",
                                  "leftThirdQuantity", "leftFourthProd", "leftFourthQuantity", "rightFirstProd",
                                  "rightFirstQuantity",
                                  "rightSecondProd", "rightSecondQuantity", "rightThirdProd", "rightThirdQuantity",
                                  "rightFourthProd",
                                  "rightFourthQuantity", "paperNumber", "cargoTrackingNumberNumeric", "fullAddress"]
        self.waiting_orders = waiting_orders
        self.initUI()

    def initUI(self):
        self.setWindowTitle("BEKLEYEN SİPARİŞLER")
        self.showMaximized()

        self.select_all_button.clicked.connect(self.select_all_cbox)
        self.deselect_all_button.clicked.connect(self.deselect_all_cbox)

        self.extract_orders_button.setIcon(self.extract_xlsx_icon)
        self.extract_orders_button.setEnabled(False)  # Başlangıçta butonu pasif yap

        self.table.setColumnCount(len(self.column_list))
        self.table.setHorizontalHeaderLabels(self.column_list)

        self.extract_orders_button.clicked.connect(self.export_to_excel)

        if os.path.exists(os.path.join(database_directory_name, "label_datas.db")):
            row_len = self.waiting_orders.shape[0]
            self.table.setRowCount(row_len)
            check_box_location = len(self.column_list) - 1

            # Add filter widgets
            filter_columns = [0, 1, 2, 3,
                              4]  # Sipariş Numarası, Kargo Takip Numarası, Kargo Firması, Müşteri Adı, Soyadı
            header_layout = QHBoxLayout()
            for col in range(self.table.columnCount()):
                if col in filter_columns:
                    filter_widget = QLineEdit()
                    filter_widget.setPlaceholderText(f"Filtrele {self.column_list[col]}")
                    filter_widget.textChanged.connect(self.apply_filter)
                    header_layout.addWidget(filter_widget)
                    self.filter_widgets.append((col, filter_widget))
                else:
                    spacer = QLabel("")
                    header_layout.addWidget(spacer)

            self.ver_layout_1.addLayout(header_layout)

            for row in range(row_len):
                checkbox = QCheckBox()
                checkbox.stateChanged.connect(self.update_extract_button_state)  # Checkbox durum değişikliğine bağla
                self.table.setCellWidget(row, check_box_location, checkbox)

            self.table.resizeColumnsToContents()
            self.ver_layout_1.addWidget(self.table)
            self.ver_layout_2.addWidget(self.extract_orders_button)
            self.ver_layout_2.addWidget(self.select_all_button)
            self.ver_layout_2.addWidget(self.deselect_all_button)

            self.main_layout.addLayout(self.ver_layout_1)
            self.main_layout.addLayout(self.ver_layout_2)

            self.setLayout(self.main_layout)

    def update_extract_button_state(self):
        """Checkboxlardan herhangi biri seçiliyse butonu aktif yap, değilse pasif yap."""
        any_checked = any(self.table.cellWidget(row, len(self.column_list) - 1).isChecked()
                          for row in range(self.table.rowCount()))
        self.extract_orders_button.setEnabled(any_checked)

    @staticmethod
    def get_order_numbers_from_excel(file_name):
        try:
            workbook = load_workbook(filename=file_name)
            sheet = workbook.active

            order_numbers = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                order_number = row[0]
                if order_number is not None:
                    order_numbers.append(order_number)

        except Exception as e:
            print(e)
            return []
        else:
            return order_numbers

    def load_data_to_table(self, progress_callback=None):
        df = self.waiting_orders

        for row in range(df.shape[0]):
            for col in range(df.shape[1] - 1):
                self.table.setItem(row, col, QTableWidgetItem(str(df.iloc[row, col])))

    def select_excel_file_extract_order(self):
        try:
            file_dialog = QFileDialog()
            file_dialog.setDefaultSuffix('xlsx')

            self.excel_file_path_for_extract_orders, _ = file_dialog.getSaveFileName(self, "Excel Dosyasını Kaydet", "",
                                                                                     "Excel Dosyaları (*.xlsx);;All Files (*)")

        except Exception as e:
            print(e)

    def export_to_excel(self):
        try:
            workbook = Workbook()
            sheet = workbook.active

            sheet.append(self.excel_column_list)

            self.select_excel_file_extract_order()
            row_idx = 2
            if self.excel_file_path_for_extract_orders:
                for row in range(self.table.rowCount()):
                    checkbox = self.table.cellWidget(row, len(self.column_list) - 1)

                    if checkbox and checkbox.isChecked():
                        col_idx = 1
                        for col in range(self.table.columnCount() - 1):
                            item = self.table.item(row, col)
                            if item:
                                sheet.cell(row=row_idx, column=col_idx, value=item.text())
                                col_idx += 1
                            else:
                                print("item yok")
                        row_idx += 1

                workbook.save(self.excel_file_path_for_extract_orders)

                printed_order_numbers = self.get_order_numbers_from_excel(self.excel_file_path_for_extract_orders)

                for order_number in printed_order_numbers:
                    update_label_printing_status(order_number, "True")

                QMessageBox.information(self, 'Başarılı', 'Veriler Excel dosyasına döküldü: {}'
                                        .format(self.excel_file_path_for_extract_orders), QMessageBox.StandardButton.Ok)

        except Exception as save_except:
            QMessageBox.critical(self, 'Hata', 'Excel dosyasına dökme hatası:\n{}'.format(str(save_except)),
                                 QMessageBox.StandardButton.Ok)

    def select_all_cbox(self):
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row):  # Sadece görünen satırlar
                checkbox = self.table.cellWidget(row, len(self.column_list) - 1)
                checkbox.setCheckState(Qt.CheckState.Checked)
        self.update_extract_button_state()

    def deselect_all_cbox(self):
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row):  # Sadece görünen satırlar
                checkbox = self.table.cellWidget(row, len(self.column_list) - 1)
                checkbox.setCheckState(Qt.CheckState.Unchecked)
        self.update_extract_button_state()

    def apply_filter(self):
        for row in range(self.table.rowCount()):
            show_row = True
            for col, filter_widget in self.filter_widgets:
                filter_text = filter_widget.text().lower()
                item_text = self.table.item(row, col).text().lower() if self.table.item(row, col) else ""
                if filter_text not in item_text:
                    show_row = False
                    break
            self.table.setRowHidden(row, not show_row)


class WaitingOrderItemWindow(QWidget):
    def __init__(self, waiting_order_items):
        super().__init__()

        self.waiting_order_items = waiting_order_items

        self.table = QTableWidget()
        self.filter_widget = QLineEdit()
        self.filter_widget.setPlaceholderText("Ürün Adına Göre Filtrele")

        self.main_layout = QHBoxLayout()
        self.ver_layout_1 = QVBoxLayout()
        self.ver_layout_2 = QVBoxLayout()

        self.select_all_button = QPushButton("TÜMÜNÜ SEÇ")
        self.deselect_all_button = QPushButton("TÜM SEÇİMİ KALDIR")

        self.column_list = ["barcode", "quantity", "productSize", "merchantSku", "productName", "productCode"]

        self.initUI()

    def initUI(self):
        self.setWindowTitle("GÖNDERİLMEYİ BEKLEYEN ÜRÜNLER")
        self.showMaximized()

        self.select_all_button.clicked.connect(self.select_all_cbox)
        self.deselect_all_button.clicked.connect(self.deselect_all_cbox)

        self.table.setColumnCount(len(self.column_list) + 1)  # Checkbox sütunu için +1
        self.column_list.append("SEÇ")
        self.table.setHorizontalHeaderLabels(self.column_list)

        self.filter_widget.textChanged.connect(self.apply_filter)
        header_layout = QHBoxLayout()
        header_layout.addWidget(self.filter_widget)
        self.ver_layout_1.addLayout(header_layout)

        self.table.resizeColumnsToContents()
        self.ver_layout_1.addWidget(self.table)
        self.ver_layout_2.addWidget(self.select_all_button)
        self.ver_layout_2.addWidget(self.deselect_all_button)

        self.main_layout.addLayout(self.ver_layout_1)
        self.main_layout.addLayout(self.ver_layout_2)

        if os.path.exists(os.path.join(database_directory_name, "orders.db")):
            row_len = self.waiting_order_items.shape[0]
            self.table.setRowCount(row_len)
            check_box_location = len(self.column_list) - 1

            for row in range(row_len):
                checkbox = QCheckBox()
                self.table.setCellWidget(row, check_box_location, checkbox)



        self.setLayout(self.main_layout)

    def load_data_to_table(self, progress_callback=None):

        try:

            df = self.waiting_order_items
            for row in range(df.shape[0]):
                for col in range(df.shape[1] - 1):
                    self.table.setItem(row, col, QTableWidgetItem(str(df.iloc[row, col])))

        except Exception as e:
            print("hata,",
                  e)

    def select_all_cbox(self):
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row):
                checkbox = self.table.cellWidget(row, len(self.column_list)-1)
                checkbox.setCheckState(Qt.CheckState.Checked)

    def deselect_all_cbox(self):
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row):
                checkbox = self.table.cellWidget(row, len(self.column_list)-1)
                checkbox.setCheckState(Qt.CheckState.Unchecked)

    def apply_filter(self):

        filter_text = self.filter_widget.text().lower()
        for row in range(self.table.rowCount()):
            item_text = self.table.item(row, self.column_list.index("productName")).text().lower() if self.table.item(
                row,
                self.column_list.index(
                    "productName")) else ""
            self.table.setRowHidden(row, filter_text not in item_text)


class MainWindow(QMainWindow):

    def __init__(self):

        super().__init__()


        self.license_check_wind = LicenseValidationWindow()

        self.license_status = 1
        self.total_product_price_label = QLabel()
        self.profit_label = QLabel()
        self.total_commission_label = QLabel()
        self.total_tax_label = QLabel()
        self.total_cargo_label = QLabel()
        self.total_otp_label = QLabel()
        self.total_sell_price_label = QLabel()

        self.setWindowTitle("OrderScout")
        self.setWindowIcon(QIcon("images/sniper_icon.png"))
        self.showMaximized()

        self.comp_api_list = []
        self.check_box_list = []
        self.completing_datas_list = []

        self.any_item = None
        self.check_box = None

        self.add_window = None
        self.waiting_order_window = None
        self.printed_order_window = None
        self.waiting_item_window = None

        self.worker_pull = None
        self.worker_calc = None
        self.worker_show_print_wind = None
        self.worker_show_wait_wind = None
        self.worker_show_wait_item_wind = None

        self.excel_file_path_for_profit = None
        self.excel_file_path_for_xlsx_stock = None

        self.thread_pool = QThreadPool()

        int_validator = QIntValidator()

        self.completer = QCompleter(self.completing_datas_list)

        toolbar = QToolBar("yapabileceğiniz işlemler")
        toolbar.setIconSize(QSize(32, 32))

        self.addToolBar(toolbar)

        company_add_action = QAction(QIcon("images/add_button.png"), "ŞİRKET EKLE", self)
        company_add_action.setStatusTip("Şirket eklemek için tıkla!")
        company_add_action.triggered.connect(self.show_comp_add_window)

        license_action = QAction(QIcon("images/license_icon.png"), "Lisans Bilgileri", self)
        license_action.setStatusTip("Lisans bilgilerini görüntülemek için tıkla!")
        license_action.triggered.connect(self.license_check_wind.show)

        update_action = QAction(QIcon("images/update_icon.png"), "GÜNCELLEME İŞLEMLERİ", self)
        update_action.setStatusTip("Güncelleme işlemleri için tıkla!")
        update_action.triggered.connect(self.update_program)

        toolbar.addAction(license_action)
        toolbar.addAction(company_add_action)
        toolbar.addAction(update_action)  # Güncelleme işlemlerini toolbar'a ekleyin

        self.run_icon = QIcon("images/circle-start-button.png")
        self.loop_icon = QIcon("images/loop-image.png")
        self.stop_icon = QIcon("images/stop-hand-icon.png")

        self.plot_widget = pg.PlotWidget()
        self.d_item_pos = pg.PlotDataItem()
        self.d_item_neg = pg.PlotDataItem()

        self.start_button = QPushButton()
        self.stop_button = QPushButton()
        self.waiting_orders_button = QPushButton("BEKLEYEN SİPARİŞLER")
        self.printed_orders_button = QPushButton("YAZDIRILAN SİPARİŞLER")

        self.import_xlsx_button = QPushButton()
        self.select_excel_button = QPushButton("SELECT EXCEL", self)
        self.match_button = QPushButton('MATCH', self)
        self.calculate_profit_button = QPushButton("CALCULATE PROFİT", self)
        self.waiting_order_items = QPushButton("GÖNDERİLMEYİ BEKLEYEN ÜRÜNLER", self)

        self.slider = QSlider()
        self.lcd = QLCDNumber()
        self.spin_box = QSpinBox()
        self.list_widget = QListWidget()

        self.circ_pbar = CircularProgress()
        self.circ_pbar_pull = CircularProgress()

        self.yes_checkbox = QCheckBox("Evet")
        self.auto_check_box = QCheckBox()

        self.cargo_label_edit = CurrencyTextEdit()
        self.cargo_bag_edit = CurrencyTextEdit()
        self.phb_edit = CurrencyTextEdit()
        self.other_prices_edit = CurrencyTextEdit()
        self.worker_price_edit = CurrencyTextEdit()
        self.package_price_plus_edit = CurrencyTextEdit()

        self.product_name_edit = QLineEdit(self)
        self.product_name_edit.setCompleter(self.completer)
        self.product_name_edit.textChanged.connect(self.product_text_changed)

        self.product_name_edit_2 = QLineEdit(self)
        self.product_name_edit_2.textChanged.connect(self.product_text_changed)
        self.product_name_edit_2.setCompleter(self.completer)

        self.product_price_edit = CurrencyTextEdit()

        self.product_stock_code = QLineEdit(self)

        self.purchase_location_edit = QLineEdit(self)
        self.purchase_location_edit.setCompleter(self.completer)
        self.purchase_location_edit.textChanged.connect(self.purchase_loc_text_changed)

        self.purchase_date_edit = QCalendarWidget(self)

        self.quantity_edit = QLineEdit(self)
        self.quantity_edit.setValidator(int_validator)
        self.stock_code_edit = QLineEdit(self)

        self.package_quantity = QLineEdit(self)
        self.package_quantity.setValidator(int_validator)

        self.advert_barcode = QLineEdit(self)

        self.tab_widget = QTabWidget(self)
        self.active_tab_index = self.tab_widget.currentIndex()
        self.tab_widget.currentChanged.connect(self.on_tab_changed)

        self.tab_pull_orders = QWidget()
        self.tab_add_stock = QWidget()
        self.tab_match_ad_stock = QWidget()
        self.tab_calculate_profit = QWidget()

        self.setCentralWidget(self.tab_widget)

        self.tab_widget.addTab(self.tab_pull_orders, "VERİ ÇEK")
        self.tab_widget.addTab(self.tab_add_stock, "STOK EKLE")
        self.tab_widget.addTab(self.tab_match_ad_stock, "STOK VE İLAN EŞLEŞTİR")
        self.tab_widget.addTab(self.tab_calculate_profit, "EXCEL İLE KAR HESAPLA")

        create_stock_datas_db()
        create_license_info_db()
        create_comp_info_db()
        create_label_datas_dbase()

        self.create_tab_pull_orders()
        self.create_tab_add_stock()
        self.create_tab_match_stock_advert()
        self.create_tab_calculate_profit()

        self.change_print_wait_button_activate()
        #self.license_checking_task()

    def create_tab_pull_orders(self):

        self.slider.setRange(0, 30)
        self.slider.setValue(0)
        self.slider.setFixedSize(50, 75)
        self.slider.valueChanged.connect(self.slider_value_printer)

        self.license_check_wind.validate_button.clicked.connect(self.validate_button_action)

        self.auto_check_box.setText("en son tarama tarihini baz alarak tara")
        self.auto_check_box.clicked.connect(self.change_spinbox_state)

        if get_last_scrap_date()[0][0] is None:
            self.auto_check_box.setEnabled(False)
        else:
            self.auto_check_box.setEnabled(True)

        self.lcd.setFixedSize(110, 100)

        self.spin_box.setRange(1, 500)

        self.license_check_wind.disconnect_button.clicked.connect(self.disconnect_license_button_action)

        self.list_widget.setFixedSize(400, 100)

        self.start_button.setFixedSize(120, 120)
        self.start_button.setIcon(self.run_icon)
        self.start_button.setIconSize(self.start_button.size())
        self.start_button.setFlat(True)
        self.start_button.setStyleSheet('''
                            QPushButton {
                                border: 0;
                                color: lightGray;
                                font-size: 1px;
                            }
                    ''')
        self.start_button.clicked.connect(self.start_button_action)

        self.stop_button.setFixedSize(120, 120)
        self.stop_button.setIcon(self.stop_icon)
        self.stop_button.setIconSize(self.stop_button.size())
        self.stop_button.setFlat(True)
        self.stop_button.setStyleSheet('''
                                    QPushButton {
                                        border: 0;
                                        color: lightGray;
                                        font-size: 1px;
                                    }
                            ''')
        self.stop_button.setVisible(False)
        self.stop_button.clicked.connect(self.stop_button_action)

        self.add_cbox_to_list_widget()

        self.printed_orders_button.setEnabled(False)
        self.waiting_orders_button.setEnabled(False)

        self.waiting_orders_button.setFixedSize(500, 200)
        self.waiting_orders_button.clicked.connect(self.show_waiting_window)

        self.printed_orders_button.setFixedSize(500, 200)
        self.printed_orders_button.clicked.connect(self.show_printed_window)

        layout_ver_mini_1 = QVBoxLayout()

        # Başlama Ayarları grup kutusunu oluştur
        start_settings_group_box = QGroupBox("Başlama Ayarları")
        layout_hor_mini_1 = QHBoxLayout()
        layout_hor_mini_1.addWidget(self.start_button)
        layout_hor_mini_1.addWidget(self.stop_button)
        layout_hor_mini_1.addWidget(self.lcd)
        layout_hor_mini_1.addWidget(self.slider)
        start_settings_group_box.setLayout(layout_hor_mini_1)

        # Aranacak Aralık grup kutusunu oluştur
        search_range_group_box = QGroupBox("Aranacak Aralık")
        layout_hor_mini_2 = QHBoxLayout()
        layout_hor_mini_2.addWidget(self.auto_check_box)
        layout_hor_mini_2.addWidget(self.spin_box)
        layout_ver_mini_1.addLayout(layout_hor_mini_2)
        layout_ver_mini_1.addWidget(self.list_widget)

        search_range_group_box.setLayout(layout_ver_mini_1)

        layout_hor_mini_3 = QHBoxLayout()

        layout_hor_main_1 = QHBoxLayout()
        layout_ver_main_1 = QVBoxLayout()
        layout_ver_main_2 = QVBoxLayout()

        layout_ver_main_1.addWidget(start_settings_group_box)
        layout_ver_main_1.addWidget(search_range_group_box)

        layout_hor_mini_3.addWidget(self.printed_orders_button)
        layout_hor_mini_3.addWidget(self.waiting_orders_button)

        layout_ver_main_2.addLayout(layout_hor_mini_3)
        layout_ver_main_2.addWidget(self.circ_pbar_pull)

        layout_hor_main_1.addLayout(layout_ver_main_1)
        layout_hor_main_1.addLayout(layout_ver_main_2)

        self.tab_pull_orders.setLayout(layout_hor_main_1)

    def create_tab_add_stock(self):
        h_layout_main = QHBoxLayout()
        v_layout_main = QVBoxLayout()

        form_layout = QFormLayout()
        save_button = QPushButton('Kaydet', self)
        save_button.clicked.connect(self.save_stock_datas_button_action)
        self.waiting_order_items.clicked.connect(self.show_waiting_item_window)
        self.import_xlsx_button.setFixedSize(120, 100)
        self.import_xlsx_button.setText("EXCEL İLE STOK EKLE")
        self.import_xlsx_button.clicked.connect(self.import_stock_datas_with_xlsx)

        # Ürün adı alanı

        form_layout.addRow('Ürün Adı:', self.product_name_edit)

        # Ürün fiyatı alanı

        form_layout.addRow('Ürün Fiyatı:', self.product_price_edit)

        # Alım yeri alanı

        form_layout.addRow('Alım Yeri:', self.purchase_location_edit)

        # Alım tarihi alanı

        form_layout.addRow('Alım Tarihi:', self.purchase_date_edit)

        # Miktarı alanı

        form_layout.addRow('Miktar:', self.quantity_edit)

        # Stok kodu alanı

        form_layout.addRow('Stok Kodu:', self.stock_code_edit)

        # Kaydet düğmesi
        form_layout.addRow("Paketleme maliyeti var mı ?(patpat vs.)", self.yes_checkbox)
        form_layout.addRow(save_button)

        v_layout_main.addWidget(self.import_xlsx_button)
        v_layout_main.addWidget(self.waiting_order_items)

        h_layout_main.addLayout(form_layout)
        h_layout_main.addLayout(v_layout_main)

        self.tab_add_stock.setLayout(h_layout_main)

    def create_tab_match_stock_advert(self):
        form_layout = QFormLayout()
        self.match_button.clicked.connect(self.match_button_action)

        # Ürün adı alanı

        form_layout.addRow('Ürün Adı:', self.product_name_edit_2)

        form_layout.addRow('Ürün stok kodu:', self.product_stock_code)
        # Miktarı alanı

        form_layout.addRow('Miktar:', self.package_quantity)

        # Stok kodu alanı

        form_layout.addRow('İlan Barkodu:', self.advert_barcode)

        # Kaydet düğmesi

        form_layout.addRow(self.match_button)
        self.package_quantity.setText("1")

        self.tab_match_ad_stock.setLayout(form_layout)

    def create_tab_calculate_profit(self):
        try:
            layout_hor_main = QHBoxLayout()
            layout_ver_main = QVBoxLayout()
            layout_form_main = QFormLayout()
            layout_form_results = QFormLayout()
            layout_hor_mini_1 = QHBoxLayout()
            layout_hor_mini_2 = QHBoxLayout()

            group_box_main = QGroupBox("EK MALİYETLER")
            group_box_results = QGroupBox("SONUÇLAR")

            self.phb_edit.setFixedSize(70, 30)
            self.cargo_bag_edit.setFixedSize(70, 30)
            self.cargo_label_edit.setFixedSize(70, 30)
            self.package_price_plus_edit.setFixedSize(70, 30)
            self.worker_price_edit.setFixedSize(70, 30)
            self.other_prices_edit.setFixedSize(70, 30)

            total_prof_row_name = QLabel("TOTAL KAR:")
            total_prod_row_name = QLabel("TOTAL ÜRÜN:")
            total_com_row_name = QLabel("TOTAL KOMİSYON:")
            total_tax_row_name = QLabel("TOTAL VERGİ:")
            total_cargo_row_name = QLabel("TOTAL KARGO:")
            total_otp_row_name = QLabel("TOTAL DİĞER:")
            total_sell_price_row_name = QLabel("TOTAL CİRO")

            total_prod_row_name.setFont(QFont("Arial", 12))
            total_prof_row_name.setFont(QFont("Arial", 12))
            total_com_row_name.setFont(QFont("Arial", 12))
            total_tax_row_name.setFont(QFont("Arial", 12))
            total_cargo_row_name.setFont(QFont("Arial", 12))
            total_otp_row_name.setFont(QFont("Arial", 12))
            total_sell_price_row_name.setFont(QFont("Arial", 12))

            layout_form_results.addRow(total_prof_row_name, self.profit_label)
            layout_form_results.addRow(total_prod_row_name, self.total_product_price_label)
            layout_form_results.addRow(total_com_row_name, self.total_commission_label)
            layout_form_results.addRow(total_tax_row_name, self.total_tax_label)
            layout_form_results.addRow(total_cargo_row_name, self.total_cargo_label)
            layout_form_results.addRow(total_otp_row_name, self.total_otp_label)
            layout_form_results.addRow(total_sell_price_row_name, self.total_sell_price_label)

            layout_form_main.addRow("platform hizmet bedeli(kdv hariç)", self.phb_edit)
            layout_form_main.addRow("kargo poşeti maliyeti(kdv hariç)", self.cargo_bag_edit)
            layout_form_main.addRow("etiket maliyeti(kdv hariç) ", self.cargo_label_edit)
            layout_form_main.addRow("ek paketleme maliyeti(kdv hariç) ", self.package_price_plus_edit)
            layout_form_main.addRow("işçilik maliyeti(kdv hariç) ", self.worker_price_edit)
            layout_form_main.addRow("diğer maliyetler(kdv hariç)", self.other_prices_edit)

            self.phb_edit.setDisabled(True)
            self.cargo_label_edit.setDisabled(True)
            self.cargo_bag_edit.setDisabled(True)
            self.package_price_plus_edit.setDisabled(True)
            self.worker_price_edit.setDisabled(True)
            self.other_prices_edit.setDisabled(True)

            layout_hor_mini_1.addWidget(self.select_excel_button)
            layout_hor_mini_1.addWidget(self.calculate_profit_button)

            group_box_results.setLayout(layout_form_results)
            layout_hor_mini_2.addWidget(self.circ_pbar)
            layout_hor_mini_2.addWidget(group_box_results)

            group_box_main.setLayout(layout_form_main)

            layout_ver_main.addLayout(layout_hor_mini_2)
            layout_ver_main.addLayout(layout_hor_mini_1)
            layout_ver_main.addWidget(group_box_main)

            layout_hor_main.addLayout(layout_ver_main)
            layout_hor_main.addWidget(self.plot_widget)

            self.total_product_price_label.setFont(QFont("Arial", 15))
            self.profit_label.setFont(QFont("Arial", 15))
            self.total_commission_label.setFont(QFont("Arial", 15))
            self.total_tax_label.setFont(QFont("Arial", 15))
            self.total_cargo_label.setFont(QFont("Arial", 15))
            self.total_otp_label.setFont(QFont("Arial", 15))
            self.total_sell_price_label.setFont(QFont("Arial", 15))

            self.plot_widget.setBackground("black")
            self.calculate_profit_button.setDisabled(True)
            self.select_excel_button.setFixedSize(120, 80)

            self.select_excel_button.clicked.connect(self.select_excel_file_profit)
            self.calculate_profit_button.clicked.connect(self.calculate_profit_button_action)
            self.tab_calculate_profit.setLayout(layout_hor_main)

        except Exception as e:
            print(e)

    @staticmethod
    def find_base_dir():
        # İçinde bulunduğumuz dizini ve üst dizinleri kontrol et
        current_dir = os.path.dirname(os.path.abspath(__file__))
        while True:
            update_script_path = os.path.join(current_dir, "update_script.exe")
            if os.path.exists(update_script_path):
                return current_dir

            # Kök dizine ulaştıysak döngüyü sonlandır
            parent_dir = os.path.dirname(current_dir)
            if parent_dir == current_dir:
                raise FileNotFoundError("update_script.exe bulunamadı")

            current_dir = parent_dir

    def update_program(self):
        check_results = update_check_tasks.check_for_updates()
        latest_version = check_results["tag_name"].strip().lstrip("v.")
        current_version = update_check_tasks.read_current_version()

        if latest_version > current_version:
            download_url = check_results["zipball_url"]

            # Mevcut ve güncellenecek sürümleri içeren bir mesaj kutusu oluştur
            message_box = QMessageBox()
            message_box.setIcon(QMessageBox.Icon.Question)
            message_box.setWindowTitle("Güncelleme Mevcut")
            message_box.setText(
                f"Mevcut sürüm: {current_version}\nGüncellenecek sürüm: {latest_version}\nGüncellemek ister misiniz?")
            message_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

            reply = message_box.exec()

            if reply == QMessageBox.StandardButton.Yes:
                try:
                    program_dir = self.find_base_dir()
                    update_script = os.path.join(program_dir, "update_script.exe")
                    print(update_script)
                    # Güncelleme scriptini çalıştır
                    subprocess.Popen(
                        [update_script, program_dir, download_url, database_directory_name,
                         latest_version])

                    # Ana programı kapat
                    QApplication.instance().quit()

                except Exception as e:
                    print(e)
            else:
                QMessageBox.information(self, "Güncelleme", "Güncelleme iptal edildi.")
        else:
            QMessageBox.information(self, "Güncelleme", "Programınız güncel.")

    def change_pbar_value_profit_tab(self, progres_value):
        self.circ_pbar.setProgress(progres_value)

    def select_excel_file_profit(self):
        try:
            file_dialog = QFileDialog(self, "Excel Dosyasını Seç", "", "Excel Dosyaları (*.xlsx);;Tüm Dosyalar (*)")

            if file_dialog.exec():
                selected_file = file_dialog.selectedFiles()
                if selected_file:
                    self.excel_file_path_for_profit = selected_file[0]
                    self.calculate_profit_button.setEnabled(True)
                    self.cargo_label_edit.setEnabled(True)
                    self.phb_edit.setEnabled(True)
                    self.cargo_bag_edit.setEnabled(True)
                    self.other_prices_edit.setEnabled(True)
                    self.package_price_plus_edit.setEnabled(True)
                    self.worker_price_edit.setEnabled(True)
        except Exception as e:
            print(e)

    def select_excel_file_xlsx_stock(self):
        try:
            file_dialog = QFileDialog(self, "Excel Dosyasını Seç", "", "Excel Dosyaları (*.xlsx);;Tüm Dosyalar (*)")

            if file_dialog.exec():
                selected_file = file_dialog.selectedFiles()
                if selected_file:
                    self.excel_file_path_for_xlsx_stock = selected_file[0]

        except Exception as e:
            print(e)

    def import_stock_datas_with_xlsx(self):
        self.select_excel_file_xlsx_stock()
        if self.excel_file_path_for_xlsx_stock:
            data_tasks.push_stock_datas_to_tuple_from_xlsx(self.excel_file_path_for_xlsx_stock)
            push_stock_datas_to_dbase(data_tasks.stock_datas_list)

    def match_button_action(self):
        self.add_checked_comp_datas_to_list()
        try:
            if (self.product_name_edit_2.text().strip() == "" or
                    self.advert_barcode.text().strip() == "" or
                    self.product_stock_code.text().strip() == "" or
                    self.package_quantity.text().strip() == ""):
                QMessageBox.critical(self, 'Hata', 'Kayıt tamamlanamadı, eksik alanlar var')

            elif len(self.comp_api_list) == 0:
                QMessageBox.critical(self, 'Hata', 'Kayıt tamamlanamadı, şirket seçiniz...')

            else:

                self.add_checked_comp_datas_to_list()
                api = Api_engine.TrendyolApi(self.comp_api_list[0][1], self.comp_api_list[0][2],
                                             self.comp_api_list[0][0])
                barcode_test = api.find_product_with_barcode(self.advert_barcode.text()).get("content")
                if len(barcode_test) > 0:

                    match_datas_list = []
                    match_datas_tuple = (
                        self.product_stock_code.text(), self.package_quantity.text(), self.advert_barcode.text())
                    match_datas_list.append(match_datas_tuple)

                    push_match_datas_to_dbase(match_datas_list)
                    QMessageBox.information(self, 'Başarılı', 'Veriler başarıyla kaydedildi',
                                            QMessageBox.StandardButton.Ok)
                    self.clear_match_stock_texts()

                else:
                    QMessageBox.warning(self, 'Başarısız', 'Pazaryerinizde {} şeklinde bir barkod bulunamadı'
                                        .format(str(self.advert_barcode.text())),
                                        QMessageBox.StandardButton.Ok)


        except Exception as e:
            print(e)

    def save_stock_datas_button_action(self):

        try:
            if (self.product_name_edit.text().strip() == "" or
                    self.quantity_edit.text().strip() == "" or
                    self.stock_code_edit.text().strip() == "" or
                    self.purchase_location_edit.text().strip() == "" or
                    self.product_price_edit.text().strip() == ""):

                QMessageBox.critical(self, 'Hata', 'Kayıt tamamlanamadı, eksik alanlar var')
            else:

                stock_datas_list = []
                stock_datas_tuple = (self.product_name_edit.text(), self.product_price_edit.text().lstrip("₺"),
                                     self.purchase_location_edit.text(),
                                     self.purchase_date_edit.selectedDate().toString("dd.MM.yy"),
                                     self.quantity_edit.text(), self.stock_code_edit.text(),
                                     self.yes_checkbox.checkState().value
                                     )
                stock_datas_list.append(stock_datas_tuple)

                push_stock_datas_to_dbase(stock_datas_list)
                QMessageBox.information(self, 'Başarılı', 'Veriler başarıyla kaydedildi',
                                        QMessageBox.StandardButton.Ok)
                self.clear_add_stock_texts()
        except Exception as e:
            print(e)

    def calculate_profit_button_action(self):
        if self.excel_file_path_for_profit:
            phb = self.phb_edit.text().lstrip("₺")
            cargo_bag = self.cargo_bag_edit.text().lstrip("₺")
            cargo_label = self.cargo_label_edit.text().lstrip("₺")
            package_price_plus = self.package_price_plus_edit.text().lstrip("₺")
            worker_price = self.worker_price_edit.text().lstrip("₺")
            other_price = self.worker_price_edit.text().lstrip("₺")
            other_prices = float(phb) + float(cargo_bag) + float(cargo_label) \
                           + float(package_price_plus) + float(worker_price) + float(other_price)

            self.worker_calc = Worker(run_all_profit_calc_tasks, excel_file_path=self.excel_file_path_for_profit,
                                      other_prices=other_prices)
            self.thread_pool.start(self.worker_calc)
            self.worker_calc.signals.finished.connect(self.scrap_thread_complete)
            self.worker_calc.signals.progress.connect(self.change_pbar_value_profit_tab)
            self.worker_calc.signals.error.connect(self.error_print)
        else:
            QMessageBox.critical(self, 'Hata', 'kar hesaplama işlemi tamamlanamadı,excel seçiniz...')

    def disconnect_license_button_action(self):
        self.license_check_wind.close()
        if len(self.license_check_wind.token_input.text()) >= 1:

            disc_response = license_check_tasks.disconnect_license(self.license_check_wind.token_input.text(),
                                                                   self.license_check_wind.mac_id_input.text())
            if disc_response.status_code == 204:

                drop_license_info_db()
                create_license_info_db()
                QMessageBox.information(self, 'Başarılı', 'Makine - Lisans ilişkisi başarıyla kesildi',
                                        QMessageBox.StandardButton.Ok)
                self.license_checking_task()

            else:
                QMessageBox.critical(self, 'Hata', 'Makine - Lisans ilişkisi kesilemedi...')

        else:
            QMessageBox.critical(self, 'Hata', 'Makine - Lisans ilişkisi kesilemedi...token girmeniz gerekiyor!')

    def validate_button_action(self):

        check_response = license_check_tasks.validate_license(self.license_check_wind.key_input.text())
        print(check_response)
        self.license_check_wind.close()

        if check_response["meta"]["valid"] == 0:
            license_check_tasks.match_machine_with_license(check_response["data"]["id"],
                                                           self.license_check_wind.token_input.text())

            check_response = license_check_tasks.validate_license(self.license_check_wind.key_input.text())

            if check_response["meta"]["valid"] == 1:

                QMessageBox.information(self, 'Bilgi', 'makine lisansla eşleştirildi\n'
                                                       'LİSANSLAMA BAŞARIYLA TAMAMLANDI!')

                push_license_datas_to_dbase([(self.license_check_wind.key_input.text(),
                                              self.license_check_wind.token_input.text())])
                self.license_status = 1
                self.license_checking_task()


            else:
                QMessageBox.critical(self, 'Hata', 'Lisanslama işlemi tamamlanamadı\n'
                                                   'MAKİNE EŞLEŞTİRİLEMEDİ')

        else:
            self.license_status = 1
            push_license_datas_to_dbase([(self.license_check_wind.key_input.text(),
                                          self.license_check_wind.token_input.text())])
            QMessageBox.information(self, 'Bilgi', 'Lisanslama işlemi tamamlandı...')
            self.license_checking_task()


    def stop_button_action(self):
        """
        stop butona tıklandığında devreye girer
        :return: stop butonu devre dışı hale getirir ve bir dizi fonksiyon çağırır
        """
        self.worker_pull.signals.progress.connect(self.stop_scrap)
        self.stop_button.setEnabled(False)

    def start_button_action(self):
        """
        start butonuna tıklandığında çalışır
        :return: sinyallere bağlanır ve bazı butonları ve elementleri devre dışı bırakır
        """

        if self.license_status == 1:

            self.add_checked_comp_datas_to_list()
            if len(self.comp_api_list) == 0:
                QMessageBox.critical(self, 'Hata', 'Veri çekme işlemi tamamlanamadı, şirket seçiniz...')
            else:

                self.circ_pbar_pull.setProgress(0)

                self.start_button.setVisible(False)

                self.stop_button.setVisible(True)

                self.slider.setEnabled(False)

                self.list_widget.setDisabled(True)

                self.auto_check_box.setDisabled(True)

                self.spin_box.setDisabled(True)

                self.worker_pull = Worker(run_all_pull_tasks, comp_api_account_list=self.comp_api_list)
                self.thread_pool.start(self.worker_pull)

                self.worker_pull.signals.finished.connect(self.scrap_thread_complete)
                self.worker_pull.signals.progress.connect(self.change_pbar_value)
                self.worker_pull.signals.error.connect(self.error_print)
        else:
            self.license_check_wind.show()

    def license_checking_task(self):
        if os.path.exists(os.path.join(database_directory_name, "license_info.db")):

            license_status = get_license_status_from_dbase()
            license_status = license_status.fetchall()

            if len(license_status) == 1:

                check_response = license_check_tasks.validate_license(license_status[0][0])

                if check_response["meta"]["valid"] == 0:
                    self.license_status = 0
                    self.license_check_wind.validate_options_button.setVisible(True)
                    self.license_check_wind.disconnect_options_button.setVisible(False)

                    self.license_check_wind.disconnect_button.setVisible(False)
                    self.license_check_wind.key_input.setVisible(False)
                    self.license_check_wind.token_input.setVisible(False)
                    self.license_check_wind.key_label.setVisible(False)
                    self.license_check_wind.token_label.setVisible(False)

                    self.license_check_wind.table_widget.setRowCount(0)


                else:

                    self.license_status = 1

                    self.refresh_mac_table(license_status[0][0],
                                           license_status[0][1])

                    self.license_check_wind.validate_options_button.setVisible(False)
                    self.license_check_wind.disconnect_button.setVisible(True)
                    self.license_check_wind.disconnect_options_button.setVisible(True)
                    self.license_check_wind.key_input.setVisible(False)
                    self.license_check_wind.token_input.setVisible(False)
                    self.license_check_wind.key_label.setVisible(False)
                    self.license_check_wind.token_label.setVisible(False)
                    self.license_check_wind.validate_button.setVisible(False)
                    self.license_check_wind.disconnect_button.setVisible(False)




            elif len(license_status) > 1:
                drop_license_info_db()
                create_license_info_db()


            else:
                self.license_status = 0
                self.license_check_wind.validate_options_button.setVisible(True)
                self.license_check_wind.disconnect_options_button.setVisible(False)
                self.license_check_wind.mac_id_label.setVisible(False)
                self.license_check_wind.key_input.setVisible(False)
                self.license_check_wind.mac_id_input.setVisible(False)
                self.license_check_wind.token_input.setVisible(False)
                self.license_check_wind.key_label.setVisible(False)
                self.license_check_wind.token_label.setVisible(False)
                self.license_check_wind.disconnect_button.setVisible(False)

                self.license_check_wind.table_widget.setRowCount(0)



        else:
            QMessageBox.critical(self, 'Hata', 'Lisanslama başarısız\n'
                                               'Lisanslara ilişkin veritabanı bulunamadı,programı yeniden yükleyiniz!')

    def clear_add_stock_texts(self):
        self.product_name_edit.clear()
        self.stock_code_edit.clear()
        self.purchase_location_edit.clear()
        self.product_price_edit.clear()
        self.quantity_edit.clear()

    def clear_match_stock_texts(self):
        self.product_name_edit_2.clear()
        self.package_quantity.clear()
        self.product_stock_code.clear()
        self.advert_barcode.clear()

    def product_text_changed(self, text):

        self.completing_datas_list.clear()
        data = get_product_name_from_dbase(text).fetchall()
        if self.active_tab_index == 1:
            if len(self.product_name_edit.text()) == 0 or len(data) == 0:

                self.stock_code_edit.clear()
                self.stock_code_edit.setEnabled(True)
            else:
                self.complete_stock_code(self.product_name_edit.text())
        elif self.active_tab_index == 2:

            if len(self.product_name_edit_2.text()) == 0 or len(data) == 0:
                self.product_stock_code.clear()
                self.product_stock_code.setEnabled(True)
            else:
                self.complete_stock_code(self.product_name_edit_2.text())

        for i in data:
            self.completing_datas_list.append(i)
        try:
            item_list = [QStandardItem(item[0]) for item in self.completing_datas_list]
            model = QStandardItemModel()
            model.appendRow(item_list)
            self.completer.setModel(model)

        except Exception as err_text:
            print(err_text)

    def on_tab_changed(self, tab_index):
        self.active_tab_index = tab_index

    def purchase_loc_text_changed(self, text):
        self.completing_datas_list.clear()
        data = get_purchase_place_from_dbase(text).fetchall()

        if len(data) > 0:
            for i in data:
                self.completing_datas_list.append(i)
            try:
                item_list = [QStandardItem(item[2]) for item in self.completing_datas_list]
                model = QStandardItemModel()
                model.appendRow(item_list)
                self.completer.setModel(model)

            except Exception as err_text:
                print(err_text)
        else:
            pass

    def add_cbox_to_list_widget(self):
        comp_datas = get_comp_datas_from_dbase()
        for comp_data in comp_datas:
            any_item = QListWidgetItem(comp_data[3])
            any_item.setFlags(any_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            any_item.setCheckState(Qt.CheckState.Unchecked)
            self.list_widget.addItem(any_item)
            self.check_box_list.append(any_item)

    def add_cbox_check_states_to_dict(self):
        check_dict = {}
        for i in self.check_box_list:
            check_dict[i.text()] = str(i.checkState())

        return check_dict

    def add_checked_comp_datas_to_list(self):
        self.comp_api_list.clear()
        for cbox_state in self.add_cbox_check_states_to_dict().items():
            if cbox_state[1] == "CheckState.Checked":
                comp_data = get_selected_comps_datas_from_dbase(cbox_state[0])[0]
                self.comp_api_list.append(comp_data)
            else:
                pass

    @staticmethod
    def error_print(error):
        print(error)

    def scrap_thread_complete(self):
        """
        thread ile birlikte işlem bitince devreye girer
        :return: bazı butonları deactive hale getirir ve çıkma sırasını normale döndürür
        """
        self.stop_button.setVisible(False)
        self.stop_button.setEnabled(True)

        self.start_button.setEnabled(True)
        self.start_button.setVisible(True)

        self.slider.setEnabled(True)
        self.list_widget.setEnabled(True)

        self.auto_check_box.setEnabled(True)

        quit_turn[0] = 1

    @staticmethod
    def stop_scrap(progress):
        """
        işlem durdurulmak istendiğinde çalışır
        :param progress: bir slot görevi görür ve aldığı sinyal işlenir
        :return: progress parametresi 100 olursa çıkış sırası 0 değerine çekilir ve çıkmak için beklenir
        """

        if progress == 100:
            quit_turn[0] = 0
        else:
            pass

    def show_comp_add_window(self, check):
        try:

            if self.add_window is None:
                self.add_window = CompAddWindow()
            self.add_window.show()

        except Exception as e:
            print(check, e)

    def show_waiting_window(self, check):

        try:

            self.waiting_order_window = WaitingOrderWindow(get_ready_order_labels())

            self.worker_show_wait_wind = Worker(self.waiting_order_window.load_data_to_table)
            self.worker_show_wait_wind.signals.error.connect(self.error_print)
            self.thread_pool.start(self.worker_show_wait_wind)

            self.waiting_order_window.show()

        except Exception as e:
            print(check, e)

    def show_printed_window(self, check):
        try:

            self.printed_order_window = PrintedOrderWindow(get_printed_order_labels())

            self.worker_show_print_wind = Worker(self.printed_order_window.load_data_to_table)
            self.worker_show_print_wind.signals.error.connect(self.error_print)
            self.thread_pool.start(self.worker_show_print_wind)

            self.printed_order_window.show()


        except Exception as e:
            print(check, e)

    def show_waiting_item_window(self, check):

        try:
            self.waiting_item_window = WaitingOrderItemWindow(get_waiting_order_items())

            self.worker_show_wait_item_wind = Worker(self.waiting_item_window.load_data_to_table)
            self.worker_show_wait_item_wind.signals.error.connect(self.error_print)
            self.thread_pool.start(self.worker_show_wait_item_wind)
            self.waiting_item_window.show()

        except Exception as e:
            print(check, e)

    def change_print_wait_button_activate(self):
        if get_ready_order_labels().shape[0] > 0 and get_printed_order_labels().shape[0] > 0:
            self.waiting_orders_button.setEnabled(True)
            self.printed_orders_button.setEnabled(True)

        elif get_ready_order_labels().shape[0] > 0 and get_printed_order_labels().shape[0] < 1:
            self.waiting_orders_button.setEnabled(True)
            self.printed_orders_button.setEnabled(False)

        elif get_printed_order_labels().shape[0] > 0 and get_ready_order_labels().shape[0] < 1:
            self.printed_orders_button.setEnabled(True)
            self.waiting_orders_button.setEnabled(False)

        else:
            self.printed_orders_button.setEnabled(False)
            self.waiting_orders_button.setEnabled(False)

    def slider_value_printer(self, slider_value):
        """
        slider elementinin hareketlerine göre üretilen sayıları yakalar
        :param slider_value: üretilen sayıları yakalar ve tutar
        :return: yakalanan sayılar lcd display elementine gerçek zamanlı olarak yazılır ve icon elementi değişir
        """
        if slider_value == 0:

            self.start_button.setIcon(self.run_icon)
            self.lcd.display(slider_value)

        else:
            self.start_button.setIcon(self.loop_icon)
            self.lcd.display(slider_value)

    def change_pbar_value(self, pbar_value):
        """
        progress bar elementinin üzerine değer yazar
        :param pbar_value: slot görevi görür ve gelen progress durumunu tutar
        :return: pbar_value değişkeninin tuttuğu değeri progress_bar'a yazdırır
        """
        self.circ_pbar_pull.setProgress(pbar_value)
        self.circ_pbar_pull.add_text_to_center(str(pbar_value))

    def refresh_mac_table(self, license_key, api_token):

        i = license_check_tasks.get_machines_for_license(license_key,
                                                         f"{api_token}")

        self.license_check_wind.table_widget.setRowCount(len(i))

        cnt = 0
        for machines in i:
            self.license_check_wind.table_widget.setItem(cnt, 0, QTableWidgetItem(
                machines["attributes"]["fingerprint"]))
            self.license_check_wind.table_widget.setItem(cnt, 1,
                                                         QTableWidgetItem(machines["attributes"]["name"]))
            self.license_check_wind.table_widget.setItem(cnt, 2,
                                                         QTableWidgetItem(
                                                             machines["attributes"]["platform"]))
            self.license_check_wind.table_widget.setItem(cnt, 3, QTableWidgetItem(machines["id"]))

            cnt += 1

    def change_spinbox_state(self, check_state):
        """
        eğer check_box seçilmişse spin box elementi disable eder tam tersi ise enable eder
        :param check_state: check box elementinin tıklanınca yaydığı sinyali tutar
        :return: yakalanan sinyal neticesinde iki elementi birbirine zıt statülerde tutar
        """
        if check_state:

            self.spin_box.setEnabled(False)
        else:
            self.spin_box.setEnabled(True)

    def change_p_bar_text_profit(self, text: str):
        self.circ_pbar.add_text_to_center(text)

    def complete_stock_code(self, selected_completion):
        prod_data = get_product_name_from_dbase(selected_completion).fetchall()
        if self.active_tab_index == 1:

            if len(prod_data) >= 1:
                self.stock_code_edit.setText(prod_data[0][5])
            else:
                pass
            self.stock_code_edit.setDisabled(True)
        elif self.active_tab_index == 2:
            if len(prod_data) >= 1:
                self.product_stock_code.setText(prod_data[0][5])
            else:
                pass
            self.product_stock_code.setDisabled(True)
        else:
            pass


app = QApplication(sys.argv)
window = MainWindow()
window.show()

app.exec()
